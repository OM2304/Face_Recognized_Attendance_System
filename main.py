# =================================================================
# PROJECT: NEURAL-SCAN BIOMETRIC ACCESS TERMINAL (v9.0)
# AUTHOR: OM (Refined by Gemini)
# PURPOSE: B.Tech Final Project - AI Attendance & Analytics
# =================================================================

import time
import sys
import os
import re
import pickle
import cv2
import numpy as np
import pandas as pd
from datetime import datetime
from PIL import Image
from openpyxl import Workbook, load_workbook
import customtkinter as ctk
from tkinter import filedialog
import face_recognition
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# --- [ PYINSTALLER SPLASH HANDLER ] ---
try:
    import pyi_splash
except ImportError:
    pyi_splash = None  # This was the typo! It should be None, not another import.

    
# --- [ RESOURCE PATH UTILITY ] ---
def resource_path(relative_path):
    """
    Handles absolute paths for both development and PyInstaller environments.
    - Used for AI models and internal assets.
    - NOT used for data persistence (Excel/PKL) to ensure data is writable.
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- [ GLOBAL DESIGN CONFIG - MODERN CLEAN THEME ] ---
BG_MAIN       = "#11131A"  # Deep slate background
BG_SIDEBAR    = "#1A1D24"  # Slightly lighter sidebar
BG_CARD       = "#252A36"  # Card background for vault/inputs
ACCENT_BLUE   = "#3B82F6"  # Modern bright blue
ACCENT_HOVER  = "#2563EB"
TEXT_MAIN     = "#F3F4F6"  # Crisp off-white
TEXT_MUTED    = "#9CA3AF"  # Soft gray for secondary text
COLOR_SUCCESS = "#10B981"  # Emerald green
COLOR_ERROR   = "#EF4444"  # Soft red

FONT_MAIN     = ("Roboto", 14)          # Clean sans-serif for UI
FONT_TITLE    = ("Roboto", 20, "bold")  # Larger titles
FONT_TERMINAL = ("Consolas", 12)        # Kept for logs and raw data
RADIUS_UI     = 8                       # Softer, modern corners
SIDE_W        = 340
GAP_S, GAP_M, GAP_L = 10, 20, 30

ctk.set_appearance_mode("Dark")

# =================================================================
# MODULE 1: MAIN APPLICATION CLASS
# =================================================================

class FaceApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # --- WINDOW CONFIGURATION ---
        self.title("Made with ❤️ by OM")
        self.geometry("1450x880")
        self.configure(fg_color=BG_MAIN)

        # --- DATA & ENGINE INITIALIZATION ---
        self.DB_FILE = "face_data.pkl"
        self.SUBJECTS_FILE = "subjects.pkl"
        
        self.data = self.load_data()
        self.subjects = self.load_subjects()

        # State Variables
        self.cap = None
        self.mode = "Idle"
        self.latest_rgb_frame = None
        self.last_logged_time = {}
        self.chart_canvas = None

        # --- EVENT BINDINGS ---
        self.bind("<s>", self.save_face_handler)
        self.bind("<S>", self.save_face_handler)

        # --- BASE GRID LAYOUT ---
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Build UI Sections
        self._build_sidebar()
        self._build_tabview()
        self._build_statusbar()

        # Apply Aesthetic Layer
        self.apply_styles()
        
        # Launch Optical Core
        self.start_camera()

    # --- [ UI BUILDERS ] ---

    def _build_sidebar(self):
        """Constructs the primary control column."""
        self.sidebar_frame = ctk.CTkFrame(self, width=SIDE_W, corner_radius=0, fg_color=BG_SIDEBAR)
        self.sidebar_frame.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(10, weight=1) # Spacer to push exit button down

        # 1. Branding
        self.logo_label = ctk.CTkLabel(
            self.sidebar_frame, text="NeuralScan",
            font=FONT_TITLE, text_color=ACCENT_BLUE
        )
        self.logo_label.grid(row=0, column=0, padx=GAP_M, pady=(GAP_L, GAP_M), sticky="w")

        # 2. Registration Command
        self.reg_toggle_btn = ctk.CTkButton(
            self.sidebar_frame, text="+ Add New Student",
            height=40, font=("Roboto", 13, "bold"),
            fg_color=BG_CARD, text_color=TEXT_MAIN, hover_color=BG_MAIN,
            border_width=1, border_color=ACCENT_BLUE,
            command=self.toggle_registration_form
        )
        self.reg_toggle_btn.grid(row=1, column=0, padx=GAP_M, pady=GAP_S, sticky="ew")

        # 3. Instruction Panel (Dynamic User Feedback)
        self.instruction_frame = ctk.CTkFrame(
            self.sidebar_frame, fg_color=BG_CARD, corner_radius=RADIUS_UI
        )
        self.instruction_frame.grid(row=2, column=0, padx=GAP_M, pady=(GAP_S, GAP_M), sticky="ew")
        
        self.instruction_label = ctk.CTkLabel(
            self.instruction_frame,
            text="System Online\nSelect a protocol to begin.",
            font=FONT_MAIN, text_color=TEXT_MUTED,
            wraplength=SIDE_W - 60, justify="center"
        )
        self.instruction_label.pack(padx=15, pady=15)

        # 4. Hidden Registration Interface (Toggled)
        self.registration_frame = ctk.CTkFrame(self.sidebar_frame, fg_color=BG_CARD, corner_radius=RADIUS_UI)
        self.name_entry = ctk.CTkEntry(self.registration_frame, placeholder_text="Student Name", height=38, border_width=0, fg_color=BG_MAIN)
        self.name_entry.pack(padx=GAP_M, pady=(GAP_M, GAP_S), fill="x")
        self.roll_entry = ctk.CTkEntry(self.registration_frame, placeholder_text="Roll Number", height=38, border_width=0, fg_color=BG_MAIN)
        self.roll_entry.pack(padx=GAP_M, pady=(0, GAP_S), fill="x")
        self.confirm_reg_btn = ctk.CTkButton(
            self.registration_frame, text="Scan Face (Press S)",
            fg_color=ACCENT_BLUE, hover_color=ACCENT_HOVER, text_color="white",
            font=("Roboto", 13, "bold"), height=38,
            command=self.validate_and_start
        )
        self.confirm_reg_btn.pack(padx=GAP_M, pady=(GAP_S, GAP_M), fill="x")

        # 5. Subject Configuration
        self.sub_label = ctk.CTkLabel(self.sidebar_frame, text="Select Subject", font=("Roboto", 12, "bold"), text_color=TEXT_MUTED)
        self.sub_label.grid(row=4, column=0, padx=GAP_M, pady=(GAP_L, 4), sticky="w")
        
        self.subject_menu = ctk.CTkComboBox(
            self.sidebar_frame, values=self.subjects, height=38,
            fg_color=BG_CARD, border_width=0, button_color=BG_CARD
        )
        self.subject_menu.grid(row=5, column=0, padx=GAP_M, pady=(0, GAP_S), sticky="ew")
        if self.subjects: self.subject_menu.set(self.subjects[0])
        
        self.new_sub_btn = ctk.CTkButton(
            self.sidebar_frame, text="+ Add Subject", height=32, 
            fg_color="transparent", text_color=ACCENT_BLUE, hover_color=BG_CARD,
            command=self.add_new_subject
        )
        self.new_sub_btn.grid(row=6, column=0, padx=GAP_M, pady=(0, GAP_M), sticky="ew")

        # 6. Global Action
        self.recog_btn = ctk.CTkButton(
            self.sidebar_frame, text="Start Attendance Scan", 
            height=54, fg_color=ACCENT_BLUE, hover_color=ACCENT_HOVER, text_color="white",
            font=("Roboto", 15, "bold"), command=self.start_recognition
        )
        self.recog_btn.grid(row=8, column=0, padx=GAP_M, pady=(GAP_L, GAP_S), sticky="ew")

        self.exit_btn = ctk.CTkButton(
            self.sidebar_frame, text="Exit System", height=38, 
            fg_color="transparent", border_width=1, border_color=TEXT_MUTED, text_color=TEXT_MUTED,
            hover_color=COLOR_ERROR, command=self.quit
        )
        self.exit_btn.grid(row=9, column=0, padx=GAP_M, pady=(0, GAP_L), sticky="ew")

    def _build_tabview(self):
        """Creates the main workspace navigation."""
        self.tabview = ctk.CTkTabview(
            self, segmented_button_selected_color=ACCENT_BLUE,
            segmented_button_selected_hover_color=ACCENT_HOVER,
            segmented_button_unselected_color=BG_SIDEBAR,
            fg_color=BG_MAIN, text_color=TEXT_MAIN
        )
        self.tabview.grid(row=0, column=1, padx=GAP_L, pady=(GAP_S, 0), sticky="nsew")
        self.tabview.add("Live Scan")
        self.tabview.add("Student Database")
        self.tabview.add("Analytics Dashboard")

        # --- TAB: OPTICAL SCAN ---
        self.scan_tab = self.tabview.tab("Live Scan")
        self.scan_tab.grid_columnconfigure(0, weight=7)
        self.scan_tab.grid_columnconfigure(1, weight=3)
        self.scan_tab.grid_rowconfigure(0, weight=1)
        
        # Video Container
        self.video_container = ctk.CTkFrame(self.scan_tab, fg_color=BG_SIDEBAR, corner_radius=RADIUS_UI)
        self.video_container.grid(row=0, column=0, padx=(0, GAP_S), pady=GAP_M, sticky="nsew")
        self.video_label = ctk.CTkLabel(self.video_container, text="Camera Offline", font=FONT_TITLE, text_color=TEXT_MUTED)
        self.video_label.pack(expand=True, fill="both", padx=10, pady=10)
        
        # Log Container
        self.log_frame = ctk.CTkFrame(self.scan_tab, fg_color=BG_SIDEBAR, corner_radius=RADIUS_UI)
        self.log_frame.grid(row=0, column=1, padx=(GAP_S, 0), pady=GAP_M, sticky="nsew")
        
        log_header = ctk.CTkLabel(self.log_frame, text="Activity Log", font=("Roboto", 14, "bold"), text_color=TEXT_MAIN)
        log_header.pack(pady=(15, 5), padx=15, anchor="w")
        
        self.activity_log = ctk.CTkTextbox(
            self.log_frame, font=FONT_TERMINAL, fg_color=BG_MAIN, 
            text_color=COLOR_SUCCESS, border_width=0, corner_radius=RADIUS_UI
        )
        self.activity_log.pack(expand=True, fill="both", padx=15, pady=(0, 15))

        # --- TAB: DATA VAULT ---
        self.vault_tab = self.tabview.tab("Student Database")
        self.vault_tab.grid_columnconfigure(0, weight=1)
        self.vault_tab.grid_rowconfigure(1, weight=1)
        
        self.scrollable_vault = ctk.CTkScrollableFrame(
            self.vault_tab, fg_color=BG_SIDEBAR, corner_radius=RADIUS_UI,
            label_text="Registered Students", label_text_color=TEXT_MAIN, label_font=("Roboto", 16, "bold")
        )
        self.scrollable_vault.grid(row=1, column=0, padx=GAP_M, pady=GAP_M, sticky="nsew")

        # --- TAB: ANALYTICS ---
        self.analytics_tab = self.tabview.tab("Analytics Dashboard")
        self.analytics_tab.grid_columnconfigure(0, weight=1)
        self.analytics_tab.grid_rowconfigure(1, weight=1)
        
        self.analytics_btn = ctk.CTkButton(
            self.analytics_tab, text="Load Attendance File (.xlsx)", 
            fg_color=BG_CARD, hover_color=BG_SIDEBAR, border_width=1, border_color=ACCENT_BLUE,
            text_color=ACCENT_BLUE, font=("Roboto", 14, "bold"), height=45,
            command=self.browse_attendance_file
        )
        self.analytics_btn.grid(row=0, column=0, padx=GAP_L, pady=(GAP_L, GAP_S), sticky="ew")
        
        self.chart_frame = ctk.CTkFrame(self.analytics_tab, fg_color=BG_SIDEBAR, corner_radius=RADIUS_UI)
        self.chart_frame.grid(row=1, column=0, padx=GAP_L, pady=(GAP_S, GAP_L), sticky="nsew")

    def _build_statusbar(self):
        self.status_bar = ctk.CTkFrame(self, height=36, corner_radius=0, fg_color=BG_SIDEBAR)
        self.status_bar.grid(row=1, column=0, columnspan=2, sticky="ew")
        self.status_label = ctk.CTkLabel(
            self.status_bar, text="Status: Ready", anchor="w", 
            font=("Roboto", 12), text_color=TEXT_MUTED
        )
        self.status_label.pack(side="left", padx=GAP_L)

    # --- [ LOGIC ENGINE ] ---

    def apply_styles(self):
        """Ensures all components follow the clean styling guidelines."""
        self.refresh_vault()

    def toggle_registration_form(self):
        if self.registration_frame.winfo_viewable():
            self.registration_frame.grid_forget()
            self.reg_toggle_btn.configure(text="+ Add New Student", fg_color=BG_CARD, text_color=TEXT_MAIN)
        else:
            self.registration_frame.grid(row=3, column=0, sticky="ew", padx=GAP_M, pady=(0, GAP_M))
            self.reg_toggle_btn.configure(text="Cancel Registration", fg_color=BG_MAIN, text_color=COLOR_ERROR)

    def validate_and_start(self):
        """Regex validation for registration fields."""
        name = self.name_entry.get().strip()
        roll = self.roll_entry.get().strip()
        
        if not roll.isdigit():
            self._update_instructions("Error: Roll Number must be numeric.", COLOR_ERROR)
            return
        if not re.match(r"^[a-zA-Z\s]+$", name):
            self._update_instructions("Error: Name must contain only letters.", COLOR_ERROR)
            return
            
        # CRITICAL: Take focus off input so keyboard 'S' doesn't type into the box
        self.focus_set()
        self.mode = "Register"
        self._update_instructions(f"Ready: Look at the camera and press 'S' to capture {name}.", ACCENT_BLUE)

    def save_face_handler(self, event):
        """Captures face encoding and saves to local PKL database."""
        if self.mode == "Register" and self.latest_rgb_frame is not None:
            # Strip 'S' from strings just in case focus was lost
            name = self.name_entry.get().strip().rstrip('sS')
            roll = self.roll_entry.get().strip().rstrip('sS')
            
            self._update_instructions("Scanning face... Please keep still.", ACCENT_BLUE)
            
            boxes = face_recognition.face_locations(self.latest_rgb_frame)
            encs = face_recognition.face_encodings(self.latest_rgb_frame, boxes)
            
            if encs:
                self.data["names"].append(name)
                self.data["rolls"].append(roll)
                self.data["encodings"].append(encs[0])
                
                with open(self.DB_FILE, "wb") as f:
                    pickle.dump(self.data, f)
                
                self.refresh_vault()
                self.toggle_registration_form()
                self.name_entry.delete(0, 'end'); self.roll_entry.delete(0, 'end')
                self._update_instructions(f"Success: {name} registered successfully.", COLOR_SUCCESS)
                self.mode = "Idle"
            else:
                self._update_instructions("Error: Face not found. Try again.", COLOR_ERROR)

    def log_attendance(self, name, roll):
        """Handles Excel operations and log frequency capping."""
        subject = self.subject_menu.get()
        filename = f"{subject}_Attendance.xlsx"
        now = datetime.now()
        
        # 10-second cooldown per student to prevent spam logging
        if roll in self.last_logged_time:
            if (now - self.last_logged_time[roll]).total_seconds() < 10:
                return

        if not os.path.exists(filename):
            wb = Workbook(); ws = wb.active
            ws.append(["Date", "Roll No", "Name", "Time", "Status"])
        else:
            wb = load_workbook(filename); ws = wb.active

        ws.append([now.strftime("%Y-%m-%d"), roll, name, now.strftime("%H:%M:%S"), "Present"])
        wb.save(filename)
        self.last_logged_time[roll] = now
        
        # Update live feed
        self.activity_log.insert("0.0", f"[{now.strftime('%H:%M:%S')}] ✅ {name} - Present\n")

    def refresh_vault(self):
        """Redraws the Student Directory with individual Purge buttons."""
        for widget in self.scrollable_vault.winfo_children(): widget.destroy()
        
        for i, (name, roll) in enumerate(zip(self.data["names"], self.data["rolls"])):
            card = ctk.CTkFrame(self.scrollable_vault, fg_color=BG_CARD, height=55, corner_radius=6)
            card.pack(fill="x", pady=6, padx=10)
            
            # Roll Number
            ctk.CTkLabel(card, text=f"#{roll}", width=70, text_color=TEXT_MUTED, font=("Roboto", 13, "bold")).pack(side="left", padx=(15, 5))
            # Name
            ctk.CTkLabel(card, text=name, width=250, anchor="w", text_color=TEXT_MAIN, font=FONT_MAIN).pack(side="left")
            
            # Delete Button
            ctk.CTkButton(
                card, text="Remove", width=80, height=30, 
                fg_color="transparent", border_width=1, border_color=COLOR_ERROR, text_color=COLOR_ERROR,
                hover_color="#451515",
                command=lambda idx=i: self.delete_student(idx)
            ).pack(side="right", padx=15)

    def delete_student(self, idx):
        """Permanent removal of biometric node."""
        student_name = self.data["names"][idx]
        self.data["names"].pop(idx)
        self.data["rolls"].pop(idx)
        self.data["encodings"].pop(idx)
        with open(self.DB_FILE, "wb") as f:
            pickle.dump(self.data, f)
        self.refresh_vault()
        self._update_instructions(f"{student_name} removed from database.", TEXT_MUTED)

    def update_frame(self):
        """Core CV2 loop with automated face recognition."""
        ret, frame = self.cap.read()
        if ret:
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            self.latest_rgb_frame = rgb_frame
            
            if self.mode == "Recognize":
                # Safety Check: Empty Database protection
                if not self.data["encodings"]:
                    self._update_instructions("Error: Database empty. Add students first.", COLOR_ERROR)
                    self.mode = "Idle"
                else:
                    boxes = face_recognition.face_locations(rgb_frame)
                    encs = face_recognition.face_encodings(rgb_frame, boxes)
                    
                    for box, enc in zip(boxes, encs):
                        dist = face_recognition.face_distance(self.data["encodings"], enc)
                        match_idx = np.argmin(dist)
                        
                        if dist[match_idx] < 0.6: # Confidence Threshold
                            name, roll = self.data["names"][match_idx], self.data["rolls"][match_idx]
                            self.log_attendance(name, roll)
                            
                            t, r, b, l = box
                            # Draw cleaner bounding box using accent blue (BGR format for OpenCV)
                            cv2.rectangle(frame, (l, t), (r, b), (246, 130, 59), 2) 
                            cv2.putText(frame, name, (l, t-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (246, 130, 59), 2)

            # Display Output
            img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
            imgtk = ctk.CTkImage(light_image=img, dark_image=img, size=(720, 480))
            self.video_label.configure(image=imgtk, text="")
            
        self.after(10, self.update_frame)

    # --- [ ANALYTICS & DATA ] ---

    def browse_attendance_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Data", "*.xlsx")])
        if f:
            df = pd.read_excel(f)
            self.render_analytics(df)

    def render_analytics(self, df):
        """Generates Matplotlib visualization within the CTK frame."""
        if self.chart_canvas: self.chart_canvas.get_tk_widget().destroy()
        
        fig = Figure(figsize=(6, 4), dpi=100); ax = fig.add_subplot(111)
        # Match the background to our clean theme
        fig.patch.set_facecolor(BG_SIDEBAR); ax.set_facecolor(BG_SIDEBAR)
        
        col = "Name" if "Name" in df.columns else df.columns[0]
        counts = df.groupby(col).size()
        
        # Clean Bar Chart Styling
        ax.bar(counts.index.astype(str), counts.values, color=ACCENT_BLUE, edgecolor=BG_SIDEBAR)
        ax.tick_params(colors=TEXT_MUTED, labelsize=10)
        ax.set_title("Attendance Overview", color=TEXT_MAIN, fontsize=14, pad=15)
        
        # Clean up chart borders
        for spine in ax.spines.values(): 
            spine.set_visible(False)
        ax.spines['bottom'].set_visible(True)
        ax.spines['bottom'].set_color(TEXT_MUTED)
        
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw(); canvas.get_tk_widget().pack(expand=True, fill="both", padx=20, pady=20)
        self.chart_canvas = canvas

    # --- [ SYSTEM HELPERS ] ---

    def _update_instructions(self, text, color):
        self.instruction_label.configure(text=text, text_color=color)
        self.status_label.configure(text=f"Status: {text}", text_color=color)

    def load_data(self):
        if os.path.exists(self.DB_FILE):
            with open(self.DB_FILE, "rb") as f: return pickle.load(f)
        return {"names": [], "rolls": [], "encodings": []}

    def load_subjects(self):
        if os.path.exists(self.SUBJECTS_FILE):
            with open(self.SUBJECTS_FILE, "rb") as f: return pickle.load(f)
        return ["Python", "Machine Learning", "Cloud Computing"]

    def save_subjects(self):
        with open(self.SUBJECTS_FILE, "wb") as f: pickle.dump(self.subjects, f)

    def add_new_subject(self):
        dialog = ctk.CTkInputDialog(text="Enter new subject name:", title="Add Subject")
        new = dialog.get_input()
        if new and new not in self.subjects:
            self.subjects.append(new); self.save_subjects()
            self.subject_menu.configure(values=self.subjects)

    def start_camera(self):
        if not self.cap: self.cap = cv2.VideoCapture(0)
        self.update_frame()

    def start_recognition(self):
        if self.data["names"]: 
            self.mode = "Recognize"
            self._update_instructions("Scanner Active. Looking for faces...", COLOR_SUCCESS)
        else: 
            self._update_instructions("Error: No students registered.", COLOR_ERROR)

# =================================================================
# MODULE 2: SYSTEM BOOTLOADER (Splash & Progress)
# =================================================================

if __name__ == "__main__":
    # Create Initialization Window
    boot = ctk.CTk()
    boot.title("NeuralScan Boot")
    boot.geometry("460x220")
    boot.overrideredirect(True)
    
    # Center on screen
    sw, sh = boot.winfo_screenwidth(), boot.winfo_screenheight()
    boot.geometry(f"+{int(sw/2-230)}+{int(sh/2-110)}")
    boot.configure(fg_color=BG_MAIN)

    ctk.CTkLabel(boot, text="NeuralScan Terminal", font=("Roboto", 22, "bold"), text_color=ACCENT_BLUE).pack(pady=(40, 5))
    ctk.CTkLabel(boot, text="Initializing core components...", font=("Roboto", 12), text_color=TEXT_MUTED).pack()
    
    progress = ctk.CTkProgressBar(boot, width=360, height=10, progress_color=ACCENT_BLUE, fg_color=BG_CARD)
    progress.pack(pady=30); progress.set(0)

    def launch_sequence():
        # Simulated loading progress (actual logic loads in background)
        for i in range(1, 11):
            progress.set(i/10); boot.update(); time.sleep(0.08)
        
        # Close the PyInstaller Splash image
        if pyi_splash: pyi_splash.close()
        
        boot.destroy() # Close bootloader
        FaceApp().mainloop() # Start main engine

    boot.after(200, launch_sequence)
    boot.mainloop()